from openpyxl import load_workbook
import numpy as np
import pandas as pd
import pandas_bokeh
from bokeh.transform import linear_cmap
from bokeh.palettes import Spectral
from bokeh.io import curdoc

# 从本地读数据
def readData(fileName):
    # 打开本地工作表
    wb = load_workbook(fileName)
    worksheet = wb['Sheet']
    partList = []
    idList = []
    nameList = []
    levelList = []
    commentList = []
    timeList = []
    likeList = []
    i = 1
    for row in worksheet.rows:
        # 因为第一行存储的数据是数据种类，因此需要跳过哦第一行的row
        if (i == 1):
            i += 1
            continue
        # 分别动每行row读取到目标信息，并添加到相应的信息list
        part, name, id, level, comment, time, like = map(lambda cell: cell.value, row)
        partList.append(part)
        nameList.append(name)
        idList.append(id)
        levelList.append(level)
        commentList.append(comment)
        timeList.append(time)
        # 有一部分评论点赞数为0，存储到本地的数据为''，此处将''变换为0
        if(like == ''):
            likeList.append(0)
        else:
            likeList.append(int(like))
    return idList, nameList, levelList, commentList, timeList, likeList

# 每集评论数量条形图
def partCommentNum(df):
    # 定义输出类型
    pandas_bokeh.output_file("剧集评论数.html")
    # 设置绘制后端为pandas_bokeh
    pd.set_option('plotting.backend', 'pandas_bokeh')
    # 分组统计每集评论数
    episode_comment_num = df.groupby('剧集')['bili_id'].nunique().to_frame('评论数')
    # 设置纵坐标
    y = episode_comment_num['评论数']
    # 设置交互式输出信息（鼠标悬停数据显示）
    mapper = linear_cmap(field_name='评论数', palette=Spectral[11], low=min(y), high=max(y))
    # 绘图
    episode_bar = episode_comment_num.plot_bokeh.bar(
        ylabel="评论数量",
        title="分集评论数",
        color=mapper,
        alpha=0.8,
        legend=False
    )

# 根据日期统计评论日期和评论数量的关系
def dataCommentNum(df):
    pandas_bokeh.output_file("日期评论数.html")
    pd.set_option('plotting.backend', 'pandas_bokeh')
    # 字符截取【月-日】作为日期数据
    df['评论日期'] = df['评论时间'].str[5:10]
    # 统计每个日期评论数量情况
    date_comment_num = df.groupby('评论日期')['bili_id'].nunique().to_frame('评论数')
    # 设置日期类型
    date_comment_num.index = date_comment_num.index.astype('string')

    y = date_comment_num['评论数']
    # 设置交互式输出信息
    mapper = linear_cmap(field_name='评论数', palette=Spectral[11], low=min(y), high=max(y))
    # 绘图
    date_bar = date_comment_num.plot_bokeh.bar(
        figsize=(1400, 450),
        ylabel="评论数量",
        title="分日期评论数",
        color=mapper,
        alpha=0.8,
        legend=False,
    )

# 统计评论用户等级分布情况
def commentLevel(df):
    pandas_bokeh.output_file("评论等级分布饼状图.html")
    pd.set_option('plotting.backend', 'pandas_bokeh')
    # 分组统计不同等级用户数量
    vip_comment_num = df.groupby('会员等级').agg(用户数=('bili_id', 'nunique'))
    # 绘图
    usernum_pie = vip_comment_num.plot_bokeh.pie(
        y="用户数",
        colormap=Spectral[9],
        title="评论员VIP等级分布",
    )

# 统计每个用户平均评论次数
def peopleComment(df):
    pandas_bokeh.output_file("人均评论情况统计.html")
    pd.set_option('plotting.backend', 'pandas_bokeh')
    # 分别统计获取用户数和评论数
    vip_comment_num = df.groupby('会员等级').agg(用户数=('评论内容', 'nunique'),
                                                评论数=('bili_id', 'nunique')
                                                )
    # 根据评论内容和评论id统计不同用户人均评论次数
    vip_comment_num['人均评论数'] = round(vip_comment_num['评论数'] / vip_comment_num['用户数'], 2)
    y = vip_comment_num['人均评论数']
    # 设置交互数据显示
    mapper = linear_cmap(field_name='人均评论数', palette=Spectral[11], low=min(y), high=max(y))
    vipmean_bar = vip_comment_num.plot_bokeh.bar(
        y='人均评论数',
        ylabel="人均评论数",
        title="不同VIP用户人均评论数",
        color=mapper,
        alpha=0.8,
        legend=False
    )

# 统计评论回复字数情况
def commentLength(df):
    pandas_bokeh.output_file("评论字数情况统计.html")
    pd.set_option('plotting.backend', 'pandas_bokeh')
    # 获取评论字数字符串长度
    df['评论长度'] = df['评论内容'].str.len()
    df['评论长度'] = df['评论长度'].fillna(0).astype('int')

    # 绘图
    contentlen_hist = df.plot_bokeh.hist(
        y='评论长度',
        ylabel="评论数",
        bins=np.linspace(0, 100, 26),
        vertical_xlabel=True,
        hovertool=False,
        title="评论长度分布直方图",
        color='red',
        line_color="white",
        legend=False,
        #     normed=100,
    )

# 统计评论获赞情况
def commentLike(df):
    pandas_bokeh.output_file("评论点赞情况统计.html")
    pd.set_option('plotting.backend', 'pandas_bokeh')
    # 统计评论获赞数据
    date_comment_num = df.groupby('评论获赞')['bili_id'].nunique().to_frame('评论获赞数')

    contentlen_hist = date_comment_num.plot_bokeh.hist(
        # y='评论点赞数',
        ylabel="评论数",
        bins=np.linspace(0, 100, 26),
        vertical_xlabel=True,
        hovertool=False,
        title="评论点赞数直方图",
        color='red',
        line_color="white",
        legend=False,
        #     normed=100, # 横坐标显示上限
    )

if __name__ == '__main__':
    # 从本地获取数据进行可视化的数据支撑
    fileName = 'B站《决胜荒野》评论.xlsx'
    # 使用load_workbook方法读取数据，发现列数超出限制map(lambda cell: cell.value, row)方法限制，因此放弃了这种方法读取数据
    # partList, idList, nameList, levelList, commentList, timeList, likeList = readData(fileName)
    # 读取本地数据
    df = pd.read_excel(fileName, header=0)
    # 检查数据、输出前十行以及每列数据类型
    print(df.head(10))
    print(df.dtypes)
    # 每集评论数量统计
    partCommentNum(df)
    # 每天评论数量统计
    dataCommentNum(df)
    # 评论员会员等级分布
    commentLevel(df)
    # 人均评论次数统计
    peopleComment(df)
    # 评论字数情况统计
    commentLength(df)
    # 评论点赞情况统计
    commentLike(df)

